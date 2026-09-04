"""
bom_excel.py — Capa 3: lectura del Excel de instrucciones.

Lee la plantilla (hojas Config, Reemplazos, Inserciones, Ediciones) y arma, por
cada `nombre_archivo`, su paquete de instrucciones para el motor.

read_instructions(path) -> dict:
    {
      "4D-81-584Z-G": {
        "config":      {"codigo_bom": "BOM-005379", "agregar_bom": True,
                        "actualizar_revision": True},
        "reemplazos":  [("7H-83-081Z-X", "DWG-COMP-009309"), ...],
        "inserciones": [{"item": ..., "description": ..., ...}, ...],
        "ediciones":   [("4D-87-949Z-A", "cantidad", "5.00"), ...],
      },
      ...
    }
"""

import openpyxl

# Excel (Inserciones) -> clave que usa el engine en el dict de componente
_INSERT_MAP = {
    "item": "item", "descripcion": "description", "description": "description",
    "cantidad": "quantity", "quantity": "quantity", "uom": "uom", "rev": "rev",
    "level": "level", "op_seq": "op_seq", "item_seq": "item_seq",
    "item_status": "status", "status": "status",
    "item_category": "category", "category": "category",
}

_TRUE = {"si", "sí", "yes", "y", "true", "1", "x", "verdadero"}


def _as_bool(v, default=True):
    if v is None or str(v).strip() == "":
        return default
    return str(v).strip().lower() in _TRUE


def _clean(v):
    return None if v is None else str(v).strip()


def _header_row(ws):
    """Devuelve (índice_fila_0based, dict_columna->índice). Busca la fila que
    contiene 'nombre_archivo' (puede ser fila 1 o 2 si hay nota arriba)."""
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True)):
        cells = [(_clean(c).lower() if c is not None else "") for c in row]
        if "nombre_archivo" in cells:
            return i, {h: j for j, h in enumerate(cells) if h}
    return None, {}


def _rows_after_header(ws, hdr_i):
    for row in ws.iter_rows(min_row=hdr_i + 2, values_only=True):
        if any(c is not None and str(c).strip() != "" for c in row):
            yield row


def read_instructions(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    data = {}

    def bucket(name):
        return data.setdefault(name, {
            "config": {"codigo_bom": None, "agregar_bom": True,
                       "actualizar_revision": True,
                       "pagina_horizontal": False,
                       "revision_misma_linea": False,
                       "rev_bom": None},
            "reemplazos": [], "inserciones": [], "ediciones": [],
        })

    # ── Config ──────────────────────────────────────────────────────────────
    if "Config" in wb.sheetnames:
        ws = wb["Config"]; hi, col = _header_row(ws)
        if hi is not None:
            for row in _rows_after_header(ws, hi):
                naf = _clean(row[col["nombre_archivo"]]) if "nombre_archivo" in col else None
                if not naf:
                    continue
                b = bucket(naf)["config"]
                if "codigo_bom" in col:
                    b["codigo_bom"] = _clean(row[col["codigo_bom"]])
                if "agregar_bom" in col:
                    b["agregar_bom"] = _as_bool(row[col["agregar_bom"]])
                if "actualizar_revision" in col:
                    b["actualizar_revision"] = _as_bool(row[col["actualizar_revision"]])
                # ── Opciones nuevas (todas desactivadas si la columna no existe) ──
                if "pagina_horizontal" in col:
                    b["pagina_horizontal"] = _as_bool(row[col["pagina_horizontal"]], default=False)
                if "revision_misma_linea" in col:
                    b["revision_misma_linea"] = _as_bool(row[col["revision_misma_linea"]], default=False)
                if "rev_bom" in col:
                    b["rev_bom"] = _clean(row[col["rev_bom"]])

    # ── Reemplazos ──────────────────────────────────────────────────────────
    if "Reemplazos" in wb.sheetnames:
        ws = wb["Reemplazos"]; hi, col = _header_row(ws)
        if hi is not None:
            for row in _rows_after_header(ws, hi):
                naf = _clean(row[col["nombre_archivo"]]) if "nombre_archivo" in col else None
                old = _clean(row[col["componente_actual"]]) if "componente_actual" in col else None
                new = _clean(row[col["componente_nuevo"]]) if "componente_nuevo" in col else None
                if naf and old and new:
                    pair = (old, new)
                    lst = bucket(naf)["reemplazos"]
                    if pair not in lst:                    # dedup filas idénticas
                        lst.append(pair)

    # ── Inserciones ─────────────────────────────────────────────────────────
    if "Inserciones" in wb.sheetnames:
        ws = wb["Inserciones"]; hi, col = _header_row(ws)
        if hi is not None:
            for row in _rows_after_header(ws, hi):
                naf = _clean(row[col["nombre_archivo"]]) if "nombre_archivo" in col else None
                comp = {}
                for c, j in col.items():
                    if c in _INSERT_MAP and row[j] is not None and str(row[j]).strip():
                        comp[_INSERT_MAP[c]] = str(row[j]).strip()
                if naf and comp.get("item"):
                    bucket(naf)["inserciones"].append(comp)

    # ── Ediciones ───────────────────────────────────────────────────────────
    if "Ediciones" in wb.sheetnames:
        ws = wb["Ediciones"]; hi, col = _header_row(ws)
        if hi is not None:
            for row in _rows_after_header(ws, hi):
                naf = _clean(row[col["nombre_archivo"]]) if "nombre_archivo" in col else None
                item = _clean(row[col["item"]]) if "item" in col else None
                campo = _clean(row[col["campo"]]) if "campo" in col else None
                valor = _clean(row[col["valor_nuevo"]]) if "valor_nuevo" in col else None
                # valor_nuevo puede ir vacío: en 'rev' significa "calcular la
                # siguiente letra automáticamente" (A->B->C...).
                if naf and item and campo:
                    bucket(naf)["ediciones"].append((item, campo, valor or ""))

    return data
